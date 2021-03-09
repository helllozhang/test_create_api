import openpyxl


class Do_excel:
    def __init__(self,filename,sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        workbook = openpyxl.load_workbook(self.filename)
        res = list(workbook[self.sheetname].rows)
        da_li = []
        title = [i.value for i in res[0]]
        for itme in res[1:]:
            data = [v.value for v in itme]
            cases = dict(zip(title,data))
            da_li.append(cases)
        return da_li

    def write_excel(self,row,column,value):
        workbook = openpyxl.load_workbook(self.filename)
        res = workbook[self.sheetname]
        res.cell(row=row,column=column,value=value)
        workbook.save(self.filename)